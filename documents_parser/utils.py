import os
import zipfile
import git
import tempfile
import shutil
import os
import re
import numpy as np
from datetime import datetime
import docx 
import xlrd
import openpyxl
from .models import *
import aspose.pdf as pdf

def clean_text(text):
    
    pattern = r'[^a-zA-Z0-9\s.,!?;:\'\"()-]'
    
    # Remove all characters that do not match the pattern
    cleaned_text = re.sub(pattern, '', text)
    
    # Replace multiple spaces with a single space
    cleaned_text = re.sub(r'\s+', ' ', cleaned_text)
    
    return cleaned_text.strip()


def search_doc(docx_path,tag,file_name,user="Test_User"):    
    doc = docx.Document(docx_path)

    extracted_data_exact = []
    extracted_data_partial=[]
    current_page = 1
    # Iterate through paragraphs and store them with their paragraph number
    for para_idx, para in enumerate(doc.paragraphs):
        paragraph_text = para.text.strip()
        if paragraph_text:  
            exact_matches = list(re.finditer(rf"\b{re.escape(tag)}\b", paragraph_text, re.IGNORECASE))
            partial_matches = list(re.finditer(rf"\w*{re.escape(tag)}\w*", paragraph_text, re.IGNORECASE))
            for match in exact_matches:
                match_start = match.start()
                match_end = match.end()
                line_index, line_text =find_line_and_number_within_paragraph(paragraph_text, match_start, match_end)
                extracted_data_exact.append({
                    "Source File Name": f"{file_name}",
                    "File Type": 'Word',
                    "Tag Searched": tag,
                    "Block/Record": line_text,
                    "Location of the Tag": f"Paragraph {para_idx+1}, Line {line_index})",
                    "Date of Search": datetime.now().strftime("%B %d, %Y"),
                    "Search Author": user,
                    "Other": ""
                })
            for match in partial_matches:
                matched_text = match.group()
                if matched_text.lower() != tag.lower():  
                    match_start = match.start()
                    match_end = match.end()
        
                    line_index, line_text =find_line_and_number_within_paragraph(paragraph_text, match_start, match_end)
                    extracted_data_partial.append({
                        "Source File Name": f"{file_name}",
                        "File Type": 'Word',
                        "Tag Searched": tag,
                        "Block/Record": line_text,
                        "Location of the Tag": f"Page {current_page} (Paragraph {para_idx+1}, Line {line_index})",
                        "Date of Search": datetime.now().strftime("%B %d, %Y"),
                        "Search Author": user,
                        "Other": "partial"
                    })
    
                # Check for page break in the paragraph's properties
            if para.paragraph_format.page_break_before:
                current_page += 1


    return extracted_data_exact,extracted_data_partial 


def find_line_and_number_within_paragraph(paragraph_text, tag_start, tag_end):
    
    if not isinstance(paragraph_text, str):
        raise ValueError("The 'paragraph_text' should be of type 'str'.")
    if not isinstance(tag_start, int) or not isinstance(tag_end, int):
        raise ValueError("Both 'tag_start' and 'tag_end' should be of type 'int'.")
    if tag_start < 0 or tag_end > len(paragraph_text) or tag_start > tag_end:
        raise ValueError("Invalid 'tag_start' or 'tag_end' indices.")

    cumulative_length = 0

    lines = paragraph_text.splitlines() if '\n' in paragraph_text else paragraph_text.split('. ')

    # Iterate through lines to find the one containing the tag
    for line_idx, line in enumerate(lines):
        line_length = len(line) + 1  # +1 for the newline character

        # Check if the tag's start index falls within the current line
        if cumulative_length <= tag_start < cumulative_length + line_length:
            line_relative_start = tag_start - cumulative_length
            line_relative_end = tag_end - cumulative_length

            return line_idx + 1 , line.strip()
               
            

        cumulative_length += line_length

    return None




def extract_text_from_excel(excel_path,tag,file_name ,user="Test_User"):

    extracted_data_exact = []
    extracted_data_partial = []

    def get_column_letter(column_index):
        return openpyxl.utils.get_column_letter(column_index + 1)

    if excel_path.endswith('.xls'):
        workbook = xlrd.open_workbook(excel_path)
        for sheet in workbook.sheets():
            sheet_name = sheet.name
            print(sheet_name)
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                for col_idx, cell in enumerate(row):
                    
                    cell_text = str(cell.value)
                    
                    exact_matches = list(re.finditer(rf"\b{re.escape(tag)}\b", cell_text, re.IGNORECASE))
                    for match in exact_matches:
                        match_start = match.start()
                        match_end = match.end()
                        cell_letter = get_column_letter(col_idx)
                        extracted_data_exact.append({
                            "Source File Name": f"{file_name}",
                            "File Type": 'Excel',
                            "Tag Searched": tag,
                            "Block/Record": cell_text,
                            "Location of the Tag": f"Sheet:{sheet_name}, Row: {row_idx + 1}, Cell: {cell_letter}",
                            "Date of Search": datetime.now().strftime("%B %d, %Y"),
                            "Search Author": user,
                            "Other": ""
                        })

                    # Find partial matches in the cell text
                    partial_matches = list(re.finditer(rf"\w*{re.escape(tag)}\w*", cell_text, re.IGNORECASE))
                    for match in partial_matches:
                        matched_text = match.group()
                        if matched_text.lower() != tag.lower():  # Avoid duplicates with exact matches
                            match_start = match.start()
                            match_end = match.end()
                            cell_letter = get_column_letter(col_idx)
                            extracted_data_partial.append({
                                "Source File Name": f"{file_name}",
                                "File Type": 'Excel',
                                "Tag Searched": tag,
                                "Block/Record": cell_text,
                                "Location of the Tag": f"Sheet:{sheet_name}, Row: {row_idx + 1}, Cell: {cell_letter}",
                                "Date of Search": datetime.now().strftime("%B %d, %Y"),
                                "Search Author": user,
                                "Other": "Partial match"
                            })

    elif excel_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(excel_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for col_idx, cell in enumerate(row):
                    if cell is not None:
                        cell_text = str(cell)
                        
                        # Find exact matches in the cell text
                        exact_matches = list(re.finditer(rf"\b{re.escape(tag)}\b", cell_text, re.IGNORECASE))
                        for match in exact_matches:
                            match_start = match.start()
                            match_end = match.end()
                            cell_letter = get_column_letter(col_idx)
                            extracted_data_exact.append({
                                "Source File Name": f"{excel_path}",
                                "File Type": 'Excel',
                                "Tag Searched": tag,
                                "Block/Record": cell_text,
                                "Location of the Tag": f"Sheet:{sheet_name} Row: {row_idx}, Cell: {cell_letter}",
                                "Date of Search": datetime.now().strftime("%B %d, %Y"),
                                "Search Author": user,
                                "Other": ""
                            })

                        # Find partial matches in the cell text
                        partial_matches = list(re.finditer(rf"\w*{re.escape(tag)}\w*", cell_text, re.IGNORECASE))
                        for match in partial_matches:
                            matched_text = match.group()
                            if matched_text.lower() != tag.lower():  # Avoid duplicates with exact matches
                                match_start = match.start()
                                match_end = match.end()
                                cell_letter = get_column_letter(col_idx)
                                extracted_data_partial.append({
                                    "Source File Name": f"{excel_path}",
                                    "File Type": 'Excel',
                                    "Tag Searched": tag,
                                    "Block/Record": cell_text,
                                    "Location of the Tag": f" Sheet:{sheet_name} Row: {row_idx}, Cell: {cell_letter}",
                                    "Date of Search": datetime.now().strftime("%B %d, %Y"),
                                    "Search Author": user,
                                    "Other": "Partial match"
                                })
    return extracted_data_exact , extracted_data_partial





